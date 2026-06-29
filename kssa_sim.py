#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Faithful re-implementation (port) of TVC Pro module 4.1 (KSSA & Box-Jenkins / AR-SSA CORE)
and the minimal upstream dependencies (TVC quadratic kernel regression for t_stat/tvc,
vol_norm, dyn_ER) required to reproduce the KSSA reversal signals bar-by-bar.

Goal: run the algorithm over Polus(1).xlsx with the indicator's DEFAULT settings and
compare generated kssa_turn_up / kssa_turn_dn against the manually-marked 'Alarm' column.

All math mirrors the Pine v6 source 1:1. Eigen-decomposition uses numpy.linalg.eigh
(symmetric), which is mathematically equivalent to Pine's matrix.eigenvalues/eigenvectors
on the symmetric PSD covariance matrix; the MGS re-orthogonalisation in Pine is a no-op
on already-orthonormal eigh output, and the SSA-LRF / projector reconstruction are both
eigenvector-sign invariant.
"""
import math
import numpy as np
import openpyxl

# ----------------------------------------------------------------------------- defaults
REG_LOG          = True          # reg_mode = "Logarithmic"
H_MAX            = 9             # h_dyn (use_adapt=false)
USE_QUAD1        = True
USE_VW           = False
K_TYPE           = "Tricube"
ER_LEN           = 10

# KSSA defaults
KSSA_LEN         = 25            # L
KSSA_COMP        = 3            # fixed rank r
KSSA_BJ_LEN      = 60           # N
KSSA_AR_ORDER    = 8            # p
KSSA_FORECAST    = 15          # max_f_bars
KSSA_USE_MAD     = True
KSSA_AR_REG      = 1e-4         # alpha
KSSA_AR_TAU      = 0.5
KSSA_USE_MSSA    = False
KSSA_VOL_W       = 0.5
KSSA_USE_FADING  = True
KSSA_FADING_MAX  = 0.99
KSSA_FADING_MIN  = 0.85
KSSA_DIAG_WEIGHT = 1.0
KSSA_FORECAST_ALGO = "SSA LRF"
KSSA_USE_TOEPLITZ  = True
KSSA_RANK_METHOD   = "Shannon Entropy"
KSSA_USE_PREWHITEN = True
KSSA_USE_HUBER     = True
KSSA_USE_FBSSA     = True
KSSA_USE_REORTHO   = True
KSSA_UNBIAS_TOEP   = False
KSSA_STRICT_REFL   = True
KSSA_VERT_THRESH   = 0.9999

KSSA_SIG_PERC    = 45.0
KSSA_USE_COMBO   = True
KSSA_CURV_PERC   = 60.0
KSSA_COOLDOWN    = 4

# ----------------------------------------------------------------------------- load data
def load():
    wb = openpyxl.load_workbook('Polus(1).xlsx', data_only=True)
    ws = wb['Polus']
    O,H,L,C,V,A,DT = [],[],[],[],[],[],[]
    for r in range(2, ws.max_row+1):
        DT.append(ws.cell(r,1).value)
        O.append(float(ws.cell(r,2).value))
        H.append(float(ws.cell(r,3).value))
        L.append(float(ws.cell(r,4).value))
        C.append(float(ws.cell(r,5).value))
        V.append(float(ws.cell(r,6).value))
        A.append(ws.cell(r,7).value)
    return DT, np.array(O),np.array(H),np.array(L),np.array(C),np.array(V),A

DT,O,H,L,C,V,A = load()
n = len(C)
src = C.copy()
math_src = np.log(np.maximum(src,1e-12)) if REG_LOG else src.copy()

# ----------------------------------------------------------------------------- TA helpers (causal)
def sma(x, length):
    out = np.full(len(x), np.nan)
    for t in range(len(x)):
        if t+1 >= length:
            out[t] = np.mean(x[t-length+1:t+1])
    return out

def rma(x, length):
    out = np.full(len(x), np.nan)
    alpha = 1.0/length
    prev = np.nan
    for t in range(len(x)):
        if np.isnan(prev):
            if t+1 >= length:
                prev = np.mean(x[t-length+1:t+1]); out[t]=prev
        else:
            prev = alpha*x[t] + (1-alpha)*prev; out[t]=prev
    return out

def wma(x, length):
    out = np.full(len(x), np.nan)
    w = np.arange(1,length+1)
    for t in range(len(x)):
        if t+1>=length:
            seg = x[t-length+1:t+1]
            out[t] = np.sum(seg*w)/np.sum(w)
    return out

# ATR(9): tr then rma
tr = np.full(n, np.nan)
for t in range(n):
    if t==0:
        tr[t] = H[t]-L[t]
    else:
        tr[t] = max(H[t]-L[t], abs(H[t]-C[t-1]), abs(L[t]-C[t-1]))
atr_val = rma(tr, H_MAX)
atr_val = np.where(np.isnan(atr_val), 1.0, atr_val)

vol_sma = sma(V,50)
vol_norm = np.array([ (min(V[t]/vol_sma[t],10.0) if (not np.isnan(vol_sma[t]) and vol_sma[t]>0) else 1.0) for t in range(n)])

# dyn_ER
dyn_change = np.full(n,np.nan); dyn_vol=np.full(n,np.nan); dyn_ER_raw=np.zeros(n)
for t in range(n):
    base = src[t-ER_LEN] if t-ER_LEN>=0 else src[t]
    dyn_change[t]=abs(src[t]-base)
    s=0.0
    for k in range(ER_LEN):
        if t-k-1>=0:
            s+=abs(src[t-k]-src[t-k-1])
        else:
            s+=abs(src[t-k]-src[t-k]) if t-k>=0 else 0
    dyn_vol[t]= s if s!=0 else 1.0
    dyn_ER_raw[t] = dyn_change[t]/dyn_vol[t] if dyn_vol[t]!=0 else 0.0
dyn_ER_smoothed = wma(dyn_ER_raw,2)
dyn_ER_smoothed = np.where(np.isnan(dyn_ER_smoothed),0.0,dyn_ER_smoothed)

# ----------------------------------------------------------------------------- kernel weight
def f_get_weight(u, t):
    a=abs(u)
    if a>1.0: return 0.0
    if t=="Tricube": return (1.0-a**3)**3
    if t=="Epanechnikov": return 0.75*(1.0-a**2)
    if t=="Gaussian": return math.exp(-0.5*a**2)
    return 1.0-a

def f_calc_tvc(t_now, k_t, h, quad, apply_vw):
    """returns (final_b0, b1, t_stat)."""
    loop_h = math.ceil(h)
    p = 3 if quad else 2
    # pass 1: sum_w
    sum_w=0.0; sum_w_sq=0.0
    for i in range(0,loop_h+1):
        idx=t_now-i
        if idx<0: continue
        u=i/h; base_w=f_get_weight(u,k_t)
        vN = vol_norm[t_now-i] if (t_now-i>=0) else 1.0
        v_weight = (max(vN,0.01))**0.33 if apply_vw else 1.0
        w=base_w*v_weight
        if w>0:
            sum_w+=w; sum_w_sq+=w*w
    eff_n = (sum_w*sum_w)/sum_w_sq if sum_w_sq>0 else max(1.0,sum_w)
    if sum_w<=0: return (src[t_now],0.0,0.0)
    s0=s1=s2=s3=s4=sy0=sy1=sy2=0.0
    for i in range(0,loop_h+1):
        idx=t_now-i
        if idx<0: continue
        u=i/h; base_w=f_get_weight(u,k_t)
        v_weight=1.0
        raw_w=base_w*v_weight
        if raw_w>0:
            w=raw_w/max(sum_w,1e-10)
            val=max(src[idx],1e-12)
            y=math.log(val) if REG_LOG else val
            x=-i; x2=x*x
            s0+=w; s1+=w*x; s2+=w*x2
            if quad:
                s3+=w*x2*x; s4+=w*x2*x2; sy2+=w*y*x2
            sy0+=w*y; sy1+=w*y*x
    eps=1e-10; lam=1e-4
    s0_r=s0+lam; s2_r=s2+lam; s4_r=s4+lam
    b0=b1=b2=float('nan'); D=det=0.0
    if quad:
        D = s0_r*(s2_r*s4_r-s3*s3)-s1*(s1*s4_r-s3*s2_r)+s2_r*(s1*s3-s2_r*s2_r)
        if abs(D)>eps:
            b0=(sy0*(s2_r*s4_r-s3*s3)-sy1*(s1*s4_r-s2_r*s3)+sy2*(s1*s3-s2_r*s2_r))/D
            b1=(s0_r*(sy1*s4_r-s3*sy2)-sy0*(s1*s4_r-s2_r*s3)+s2_r*(s1*sy2-sy1*s2_r))/D
            b2=(s0_r*(s2_r*sy2-sy1*s3)-s1*(s1*sy2-sy1*s2_r)+sy0*(s1*s3-s2_r*s2_r))/D
    else:
        det=s0_r*s2_r-s1*s1
        if abs(det)>eps:
            b0=(sy0*s2_r-s1*sy1)/det; b1=(s0_r*sy1-s1*sy0)/det; b2=0.0
    if math.isnan(b0): return (src[t_now],0.0,0.0)
    tss=rss=0.0; avg_y=sy0/s0_r if s0_r!=0 else 0.0; smear_sum=0.0; resid=[]
    for i in range(0,loop_h+1):
        idx=t_now-i
        if idx<0: continue
        u=i/h; base_w=f_get_weight(u,k_t); raw_w=base_w
        if raw_w>0:
            w=raw_w/max(sum_w,1e-10)
            val=max(src[idx],1e-12); y=math.log(val) if REG_LOG else val
            x=-i; y_est=b0+b1*x+b2*x*x; err=y-y_est
            resid.append(abs(err)); rss+=w*err*err; tss+=w*(y-avg_y)**2
            if REG_LOG: smear_sum+=w*math.exp(err)
    dof=max(eff_n-p,1.0); var_est=rss/dof
    se_b1=0.0
    if quad:
        if abs(D)>eps: se_b1=math.sqrt(max(var_est*(s0_r*s4_r-s2_r*s2_r)/D,0.0))
    else:
        if abs(det)>eps: se_b1=math.sqrt(max(var_est*(s0_r/det),0.0))
    w_norm=sum_w/math.sqrt(sum_w_sq) if sum_w_sq>0 else sum_w
    se_b1=se_b1/max(w_norm,1e-9)
    t_stat= b1/se_b1 if se_b1!=0 else 0.0
    smearing_est = (smear_sum if smear_sum>0 else math.exp(var_est/2.0)) if REG_LOG else 1.0
    final_b0 = math.exp(b0)*smearing_est if REG_LOG else b0
    return (final_b0, b1, t_stat)

# precompute tvc, t_stat per bar
tvc_arr=np.zeros(n); t_stat_arr=np.zeros(n)
for t in range(n):
    fb0,b1,ts=f_calc_tvc(t,K_TYPE,H_MAX,USE_QUAD1,USE_VW)
    tvc_arr[t]=fb0; t_stat_arr[t]=ts
tvc_baseline = np.log(np.maximum(tvc_arr,1e-12)) if REG_LOG else tvc_arr.copy()

# ----------------------------------------------------------------------------- Burg AR (for completeness; default uses SSA LRF)
def get_burg_ar_weights(data, p, alpha, tau, strict_reflect):
    nn=len(data); a=[0.0]*p; ef=list(data); eb=list(data)
    var_x=0.0
    if nn>0:
        m_x=sum(data)/nn; ss=0.0
        for i in range(nn):
            dev=data[i]-m_x; dev=max(-1.5,min(1.5,dev)); ss+=dev*dev
        var_x=ss/nn
    ridge=alpha*var_x
    if nn>p and p>0:
        for m in range(1,p+1):
            num=0.0; den=0.0
            for i in range(m,nn):
                num+=ef[i]*eb[i-1]; den+=ef[i]**2+eb[i-1]**2
            den+=ridge*(nn-m)
            k=(2.0*num)/den if den>1e-12 else 0.0
            if strict_reflect:
                eps_sc=1e-3; k=(1 if k>=0 else -1)*min(abs(k),1.0-eps_sc)
            k=k*(1.0-math.exp(-((1.0-abs(k))**2)/max(tau,1e-6)))
            a[m-1]=k
            if m>1:
                a_prev=a[:]
                for i in range(0,m-1):
                    a[i]=a_prev[i]-k*a_prev[m-2-i]
            for i in range(nn-1,m-1,-1):
                ef_new=ef[i]-k*eb[i-1]; eb_new=eb[i-1]-k*ef[i]
                ef[i]=ef_new; eb[i]=eb_new
    return a

def get_ssa_lrf_weights(U_r, Lw, actual_r, vert_thresh):
    nu2=0.0; pi_vec=[0.0]*actual_r
    for i in range(actual_r):
        pv=U_r[Lw-1,i]; pi_vec[i]=pv; nu2+=pv*pv
    A=[0.0]*(Lw-1)
    denom=1.0-nu2
    safe_denom = denom if nu2<vert_thresh else max(1.0-vert_thresh,1e-9)
    coef=1.0/safe_denom
    for i in range(0,Lw-1):
        a_i=0.0
        for j in range(actual_r):
            a_i+=U_r[i,j]*pi_vec[j]
        A[Lw-2-i]=a_i*coef
    return A

# ----------------------------------------------------------------------------- KSSA core
def f_ar_ssa_0lag(t_now):
    out = math_src[t_now]
    bj_len=KSSA_BJ_LEN; L=KSSA_LEN; r=KSSA_COMP; p=KSSA_AR_ORDER
    if t_now < bj_len:   # bar_index >= bj_len
        return out
    dyn_er=dyn_ER_smoothed[t_now]
    dyn_rho = (KSSA_FADING_MIN + (KSSA_FADING_MAX-KSSA_FADING_MIN)*dyn_er) if KSSA_USE_FADING else 1.0
    # training window: hist_p[i] = src[t_now-(bj_len-1-i)]
    hist_p=np.array([ math_src[t_now-(bj_len-1-i)] for i in range(bj_len)])
    hist_v=np.array([ vol_norm[t_now-(bj_len-1-i)] for i in range(bj_len)])
    # robust centering
    if KSSA_USE_MAD:
        sp=np.sort(hist_p); sv=np.sort(hist_v); half=bj_len//2
        loc_p = (sp[half-1]+sp[half])/2.0 if bj_len%2==0 else sp[half]
        loc_v = (sv[half-1]+sv[half])/2.0 if bj_len%2==0 else sv[half]
        dp=np.sort(np.abs(hist_p-loc_p)); dv=np.sort(np.abs(hist_v-loc_v))
        mad_p=((dp[half-1]+dp[half])/2.0 if bj_len%2==0 else dp[half])*1.4826
        mad_v=((dv[half-1]+dv[half])/2.0 if bj_len%2==0 else dv[half])*1.4826
    else:
        loc_p=np.mean(hist_p); loc_v=np.mean(hist_v)
        mad_p=math.sqrt(np.mean((hist_p-loc_p)**2)); mad_v=math.sqrt(np.mean((hist_v-loc_v)**2))
    if mad_p>1e-9:
        hist_p=(hist_p-loc_p)/mad_p
    else:
        hist_p=hist_p-loc_p
    hist_v=hist_v-loc_v
    vol_scale=(mad_p/mad_v)*KSSA_VOL_W if (mad_v>1e-9 and mad_p>1e-9) else KSSA_VOL_W
    hist_v=hist_v*vol_scale
    # huber guard (covariance copies)
    hist_p_cov=hist_p.copy(); hist_v_cov=hist_v.copy()
    if KSSA_USE_HUBER:
        delta=1.345
        hist_p_cov=np.clip(hist_p_cov,-delta,delta)
        if KSSA_USE_MSSA:
            hist_v_cov=np.clip(hist_v_cov,-delta,delta)
    actual_L = L*2 if KSSA_USE_MSSA else L
    C_mat=np.zeros((actual_L,actual_L))
    if KSSA_USE_TOEPLITZ:
        R_pp=np.zeros(L)
        total_w=0.0
        for tt in range(bj_len):
            total_w += (dyn_rho**(bj_len-1-tt)) if KSSA_USE_FADING else 1.0
        win_w=np.zeros(L)
        if KSSA_UNBIAS_TOEP:
            for k in range(L):
                wk=0.0
                for tt in range(0,bj_len-k):
                    wk+= (dyn_rho**(bj_len-1-(tt+k))) if KSSA_USE_FADING else 1.0
                win_w[k]=wk
        for k in range(L):
            sum_pp=0.0
            for tt in range(0,bj_len-k):
                w=(dyn_rho**(bj_len-1-(tt+k))) if KSSA_USE_FADING else 1.0
                sum_pp+=hist_p_cov[tt]*hist_p_cov[tt+k]*w
            unbias_corr=(win_w[0]/max(win_w[k],1e-9)) if KSSA_UNBIAS_TOEP else 1.0
            R_pp[k]=(sum_pp/max(total_w,1e-9))*unbias_corr
        for ri in range(L):
            for ci in range(L):
                C_mat[ri,ci]=R_pp[abs(ri-ci)]
    else:
        K_hist=bj_len-L+1
        X_tmp=np.zeros((actual_L,K_hist))
        for j in range(K_hist):
            for i in range(L):
                w=(dyn_rho**(K_hist-1-j)) if KSSA_USE_FADING else 1.0
                X_tmp[i,j]=hist_p_cov[i+j]*math.sqrt(w)
        C_mat=X_tmp@X_tmp.T/K_hist
    # FB-SSA
    if KSSA_USE_FBSSA:
        C_fb=np.zeros_like(C_mat)
        for ri in range(actual_L):
            for ci in range(actual_L):
                C_fb[ri,ci]=(C_mat[ri,ci]+C_mat[actual_L-1-ri,actual_L-1-ci])/2.0
        C_mat=C_fb
    # eigen (symmetric)
    evals,evecs=np.linalg.eigh(C_mat)   # ascending
    # sort_idx ascending order index list (like array.sort_indices ascending)
    # We emulate: take i-th largest = position num_vals-1-i in ascending list
    num_vals=actual_L
    order_desc=list(range(num_vals-1,-1,-1))  # indices into ascending arrays, largest first
    # prewhiten noise estimate
    clean_ev=evals.copy().astype(float)
    noise_var=0.0; mp_threshold=0.0
    if KSSA_USE_PREWHITEN and actual_L>2:
        tail=[]
        for i in range(actual_L//2, actual_L):
            tail.append(max(evals[order_desc[i]],0.0))
        tail=np.sort(np.array(tail))
        noise_var=float(np.median(tail)) if len(tail)>0 else 0.0
        if KSSA_RANK_METHOD=="Marchenko-Pastur":
            gamma=actual_L/max(bj_len-actual_L+1,1.0)
            mp_threshold=noise_var*(1.0+math.sqrt(gamma))**2
    for i in range(actual_L):
        o=order_desc[i]; ev=abs(evals[o])
        if KSSA_RANK_METHOD=="Marchenko-Pastur":
            cev=ev if ev>mp_threshold else 0.0
        else:
            cev=max(0.0,ev-noise_var)
        clean_ev[o]=cev
    # rank selection
    actual_r=min(r,actual_L)
    if actual_L>2:
        if KSSA_RANK_METHOD=="Shannon Entropy":
            sum_eig=0.0
            for i in range(actual_L):
                sum_eig+=max(clean_ev[order_desc[i]],1e-12)
            total_H=0.0; shannon=[0.0]*actual_L
            for i in range(actual_L):
                ne=max(clean_ev[order_desc[i]],1e-12)/sum_eig
                h_i=-ne*math.log(ne); shannon[i]=h_i; total_H+=h_i
            cum=0.0; opt_r=1
            for i in range(actual_L):
                cum+=shannon[i]
                if cum/max(total_H,1e-12)>=0.95:
                    opt_r=i+1; break
            actual_r=max(1,opt_r)
        elif KSSA_RANK_METHOD=="Marchenko-Pastur":
            opt_r=0
            for i in range(actual_L):
                if abs(evals[order_desc[i]])>mp_threshold: opt_r+=1
            actual_r=max(1,opt_r)
        elif KSSA_RANK_METHOD=="Colored Noise MDL":
            min_mdl=1e12; opt_r=1
            for k in range(1,actual_L):
                sa=0.0; sgl=0.0; count=actual_L-k
                for i in range(k,actual_L):
                    eig=max(clean_ev[order_desc[i]],1e-12); sa+=eig; sgl+=math.log(eig)
                sa/=count; sgl/=count
                ar1=max(1.0-dyn_er,0.1)
                pen=0.5*k*(2*actual_L-k)*math.log(bj_len)*ar1
                lr=sgl-math.log(sa); mdl=-bj_len*count*lr+pen
                if mdl<min_mdl: min_mdl=mdl; opt_r=k
            actual_r=opt_r
    # U_raw
    U_raw=np.zeros((actual_L,actual_r))
    for i in range(actual_r):
        oi=order_desc[i]
        U_raw[:,i]=evecs[:,oi]
    # MGS reortho (no-op on orthonormal eigh vectors, kept for fidelity)
    if KSSA_USE_REORTHO:
        for ck in range(actual_r):
            if ck>0:
                for cj in range(ck):
                    dot=np.dot(U_raw[:,ck],U_raw[:,cj])
                    U_raw[:,ck]=U_raw[:,ck]-dot*U_raw[:,cj]
            nk=math.sqrt(np.dot(U_raw[:,ck],U_raw[:,ck]))
            if nk>1e-12: U_raw[:,ck]=U_raw[:,ck]/nk
    # U_r with prewhiten scale
    U_r=np.zeros((actual_L,actual_r))
    for i in range(actual_r):
        oi=order_desc[i]
        orig_ev=max(abs(evals[oi]),1e-12)
        scale=math.sqrt(max(clean_ev[oi]/orig_ev,0.0)) if KSSA_USE_PREWHITEN else 1.0
        U_r[:,i]=U_raw[:,i]*scale
    # AR coeffs
    alpha_dyn=KSSA_AR_REG*max(0.01,1.0-dyn_er)
    if KSSA_USE_MSSA or KSSA_FORECAST_ALGO=="Burg AR":
        ar_coeffs=get_burg_ar_weights(list(hist_p),p,alpha_dyn,KSSA_AR_TAU,KSSA_STRICT_REFL)
    else:
        ar_coeffs=get_ssa_lrf_weights(U_r,L,actual_r,KSSA_VERT_THRESH)
    # forecast embedding
    ext_p=list(hist_p); ext_v=list(hist_v)
    dyn_f_bars=int(max(1.0,round(KSSA_FORECAST*dyn_er)))
    p_len=len(ar_coeffs)
    for f in range(dyn_f_bars):
        pred_p=0.0; cl=len(ext_p)
        for k in range(1,p_len+1):
            pred_p+=ar_coeffs[k-1]*ext_p[cl-k]
        ext_p.append(pred_p)
        pred_v=ext_v[cl-1] if KSSA_USE_MSSA else 0.0
        ext_v.append(pred_v)
    # extended trajectory matrix
    N_ext=len(ext_p); K_cols=N_ext-L+1
    X_ext=np.zeros((actual_L,K_cols))
    for j in range(K_cols):
        for i in range(L):
            X_ext[i,j]=ext_p[i+j]
    X_rec=U_r@(U_r.T@X_ext)
    # Hankelization with Nuttall window
    target_t=bj_len-1; rec_val=0.0; sdw=0.0
    for i in range(L):
        j=target_t-i
        if 0<=j<K_cols:
            norm_i=float(i)/float(L-1) if L>1 else 0.0
            wn=0.3635819-0.4891775*math.cos(2*math.pi*norm_i)+0.1365995*math.cos(4*math.pi*norm_i)-0.0106411*math.cos(6*math.pi*norm_i)
            w=wn**KSSA_DIAG_WEIGHT
            rec_val+=X_rec[i,j]*w; sdw+=w
    final_scaled=rec_val/sdw if sdw>0 else 0.0
    out=loc_p+final_scaled*mad_p
    return out

# run core
kssa_val=np.array([f_ar_ssa_0lag(t) for t in range(n)])

# ----------------------------------------------------------------------------- kinematics + signals
def stdev(x,length):
    out=np.full(len(x),np.nan)
    for t in range(len(x)):
        if t+1>=length:
            seg=x[t-length+1:t+1]; out[t]=np.std(seg)  # population stdev (ta.stdev default biased)
    return out

def percentile_lin(series, length, perc):
    out=np.full(len(series),np.nan)
    for t in range(len(series)):
        if t+1>=length:
            seg=np.sort(series[t-length+1:t+1])
            rank=perc/100.0*(len(seg)-1)
            lo=int(math.floor(rank)); hi=int(math.ceil(rank)); frac=rank-lo
            out[t]=seg[lo]+(seg[hi]-seg[lo])*frac
    return out

kssa_slope=np.zeros(n)
for t in range(n):
    kssa_slope[t]=kssa_val[t]-(kssa_val[t-1] if t>0 else kssa_val[t])
kssa_curv=np.zeros(n)
for t in range(n):
    kssa_curv[t]=kssa_slope[t]-(kssa_slope[t-1] if t>0 else 0.0)

bjl=KSSA_BJ_LEN
slope_mean=sma(kssa_slope,bjl); slope_dev=stdev(kssa_slope,bjl)
curv_mean=sma(kssa_curv,bjl); curv_dev=stdev(kssa_curv,bjl)
sig_strength=np.zeros(n); curv_strength=np.zeros(n)
for t in range(n):
    sd=max(slope_dev[t],1e-6) if not np.isnan(slope_dev[t]) else 1e-6
    sm=slope_mean[t] if not np.isnan(slope_mean[t]) else 0.0
    sig_strength[t]=abs((kssa_slope[t]-sm)/sd)
    cd=max(curv_dev[t],1e-6) if not np.isnan(curv_dev[t]) else 1e-6
    cm=curv_mean[t] if not np.isnan(curv_mean[t]) else 0.0
    curv_strength[t]=abs((kssa_curv[t]-cm)/cd)
sig_thresh=percentile_lin(sig_strength,100,KSSA_SIG_PERC)
curv_thresh=percentile_lin(curv_strength,100,KSSA_CURV_PERC)

last_up=-10**9; last_dn=-10**9
buy_signals=[]; sell_signals=[]
turn_up_arr=[False]*n; turn_dn_arr=[False]*n
for t in range(n):
    st=max(sig_thresh[t],1.0) if not np.isnan(sig_thresh[t]) else 1.0
    ct=max(curv_thresh[t],1.0) if not np.isnan(curv_thresh[t]) else 1.0
    kssa_valid = sig_strength[t]>st
    curv_valid = curv_strength[t]>ct
    tvc_allow_up = t_stat_arr[t]>-2.0
    tvc_allow_dn = t_stat_arr[t]<2.0
    prev_curv = kssa_curv[t-1] if t>0 else 0.0
    prev_slope= kssa_slope[t-1] if t>0 else 0.0
    early_up = (kssa_curv[t]>0 and prev_curv<=0 and kssa_slope[t]<0 and tvc_allow_up)
    early_dn = (kssa_curv[t]<0 and prev_curv>=0 and kssa_slope[t]>0 and tvc_allow_dn)
    cross_up = (kssa_slope[t]>0 and prev_slope<=0)
    cross_dn = (kssa_slope[t]<0 and prev_slope>=0)
    if KSSA_USE_COMBO:
        raw_up=(early_up and curv_valid) or (cross_up and kssa_valid)
        raw_dn=(early_dn and curv_valid) or (cross_dn and kssa_valid)
    else:
        raw_up=cross_up and kssa_valid
        raw_dn=cross_dn and kssa_valid
    cd_up=(t-last_up)>=KSSA_COOLDOWN
    cd_dn=(t-last_dn)>=KSSA_COOLDOWN
    turn_up=raw_up and cd_up and tvc_allow_up   # mtf=true
    turn_dn=raw_dn and cd_dn and tvc_allow_dn
    if turn_up: last_up=t; buy_signals.append(t); turn_up_arr[t]=True
    if turn_dn: last_dn=t; sell_signals.append(t); turn_dn_arr[t]=True

# ----------------------------------------------------------------------------- evaluation vs Alarm
ref_buy=[i for i in range(n) if A[i]=='buy']
ref_sell=[i for i in range(n) if A[i]=='sell']

def match(ref, gen, tol):
    """match each ref to nearest gen within +/- tol bars; gen at >= ref-? we consider lag = gen-ref.
       Return matched pairs (ref,gen,lag), unmatched_ref, and used gen set."""
    used=set(); pairs=[]; unmatched=[]
    for rb in ref:
        best=None; bestd=None
        for g in gen:
            if g in used: continue
            d=abs(g-rb)
            if d<=tol and (bestd is None or d<bestd):
                bestd=d; best=g
        if best is not None:
            used.add(best); pairs.append((rb,best,best-rb))
        else:
            unmatched.append(rb)
    extra=[g for g in gen if g not in used]
    return pairs,unmatched,extra

for tol in [2,3,5]:
    pb,ub,eb=match(ref_buy,buy_signals,tol)
    ps,us,es=match(ref_sell,sell_signals,tol)
    lags=[p[2] for p in pb]+[p[2] for p in ps]
    print(f"=== TOLERANCE +/-{tol} bars ===")
    print(f" BUY : ref={len(ref_buy)} gen={len(buy_signals)} matched={len(pb)} missed={len(ub)} extra={len(eb)}")
    print(f" SELL: ref={len(ref_sell)} gen={len(sell_signals)} matched={len(ps)} missed={len(us)} extra={len(es)}")
    tot_ref=len(ref_buy)+len(ref_sell); tot_match=len(pb)+len(ps); tot_gen=len(buy_signals)+len(sell_signals)
    recall=tot_match/tot_ref if tot_ref else 0
    prec=tot_match/tot_gen if tot_gen else 0
    print(f" TOTAL: recall={recall:.3f} precision={prec:.3f} F1={(2*prec*recall/(prec+recall) if (prec+recall)>0 else 0):.3f}")
    if lags:
        print(f" Lag(gen-ref) mean={np.mean(lags):.2f} median={np.median(lags):.1f} std={np.std(lags):.2f} min={min(lags)} max={max(lags)}")
    print()

print(f"Total generated: BUY={len(buy_signals)} SELL={len(sell_signals)}  (ref BUY={len(ref_buy)} SELL={len(ref_sell)})")
print(f"Warmup bars (no core): first {KSSA_BJ_LEN} bars")
# count ref signals within warmup
ref_in_warm=sum(1 for i in ref_buy+ref_sell if i<KSSA_BJ_LEN)
print(f"Reference alarms inside warmup window (impossible to catch): {ref_in_warm}")
